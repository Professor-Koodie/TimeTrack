import io

from django.contrib.auth import get_user_model
from django.test import Client, TestCase
from django.urls import reverse
from openpyxl import load_workbook

from .models import Timesheet


class GenerateReportTests(TestCase):
    def setUp(self):
        self.client = Client()
        self.user = get_user_model().objects.create_user(
            username='reportuser',
            password='pass12345',
        )
        self.client.force_login(self.user)

    def test_report_includes_employee_month_totals(self):
        Timesheet.objects.create(
            employee='Alice',
            project_name='General Admin',
            task_description='General_Admin',
            week_number=1,
            hours_week=10,
        )
        Timesheet.objects.create(
            employee='Alice',
            project_name='General Admin',
            task_description='General_Admin',
            week_number=5,
            hours_week=12,
        )
        Timesheet.objects.create(
            employee='Bob',
            project_name='General Admin',
            task_description='General_Admin',
            week_number=9,
            hours_week=8,
        )

        response = self.client.get(reverse('generate_report'))

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(filename=io.BytesIO(response.content))
        self.assertIn('Employee Monthly Summary', workbook.sheetnames)

        monthly_sheet = workbook['Employee Monthly Summary']
        headers = [cell.value for cell in monthly_sheet[1]]
        self.assertEqual(headers[0], 'Employee')
        self.assertIn('January', headers)
        self.assertIn('February', headers)
        self.assertIn('March', headers)

        rows = list(monthly_sheet.iter_rows(values_only=True))
        row_by_employee = {row[0]: row for row in rows[1:] if row[0]}

        self.assertEqual(row_by_employee['Alice'][1], 10.0)
        self.assertEqual(row_by_employee['Alice'][2], 12.0)
        self.assertEqual(row_by_employee['Alice'][13], 22.0)
        self.assertEqual(row_by_employee['Bob'][3], 8.0)

        details_sheet = workbook['Entries']
        details_rows = list(details_sheet.iter_rows(values_only=True))
        self.assertIn('Month', details_rows[0])
        self.assertEqual(details_rows[1][4], 'January')
