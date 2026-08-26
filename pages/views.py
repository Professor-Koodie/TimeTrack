import json
import io
from pathlib import Path
from random import random
from urllib import request
from django.http import HttpResponse
from openpyxl import Workbook

from django.shortcuts import redirect, render
from django.views.generic import TemplateView
from django.contrib.auth.views import LoginView
from django.urls import reverse_lazy
from django.contrib.auth.decorators import login_required
from django.contrib.auth.views import LogoutView
from django.db.models import Max
from .models import Timesheet
from django.views.generic import UpdateView, DeleteView
from .forms import NewTimesheetForm, TimesheetForm

MONTH_NAMES = [
    'January', 'February', 'March', 'April', 'May', 'June',
    'July', 'August', 'September', 'October', 'November', 'December'
]

WEEK_MONTH_THRESHOLDS = [4, 8, 12, 16, 20, 24, 28, 32, 36, 40, 44, 48, 52] 

PROJECT_DISPLAY_MAP = {
    'DBN_Maziya': 'DBN Maziya',
    'SuperSites': 'Super Sites',
    'PT_Platform_Rectification_Gauteng': 'PT Platform Rectification Gauteng', 
    'Midway_Washaway': 'Midway Washaway', 
    'QMS': 'QMS',
    'Witbank_Level_Crossing': 'Witbank Level Crossing', 
    'Gautrain_Ballast_Investigation': 'Gautrain Ballast Investigation', 
    'Bramfontein_Yard': 'Bramfontein Yard',
    'JHB_Bram': 'JHB Bram',
    'PT_Rectification_Western_Cape': 'PT Rectification Western Cape', 
    'Richards_Bay': 'Richards Bay',
    'N3 Cato_Ridge': 'N3 Cato Ridge',
    'Vandyksdrift_Level_Crossing': 'Vandyksdrift Level Crossing',
    'General_Admin': 'General Admin',  
    'Public_Holiday': 'Public Holiday', 
    'Leave': 'Leave',   
    'Valterra_Trek_Scale': 'Valterra Trek Scale',
    'Sedra_Edilon': 'Sedra Edilon', 
    'Tsiko_LTA': 'Tsiko LTA',
    'Iron_Ore': 'Iron Ore',
    'RBM_Maintanance': 'RBM Maintanance', 
    'COTC': 'COTC', 
    'Bayvue_Rail_Yard': 'Bayvue Rail Yard',  
    'MMSEZ': "MMSEZ", 
}

#Hours planned per project
PLANNED_HOURS_MAP = {
    'DBN Maziya': 140,
    'JHB Bram': 180,
    'PT Rectification Western Cape': 160,
    'Richards Bay': 150,
    'N3 Cato Ridge': 170,
    'Vandyksdrift Level Crossing': 200,
    'General Admin': 120,
    'Public Holiday': 0,
    'Leave': 0
}

Task_Description_Display_Map = { #adited
    'Design': 'Design',
    'Draawings': 'Drawings',
    'Update_Designs': 'Update Designs',
    'Site_Visit': 'Site Visit',
    'General_Admin': 'General Admin',
    'Submission': 'Submission',
    'Client_Review': 'Client Review',
    'Report_Consolidation': 'Report Consolidation',
    'Other': 'Other', 
}   

def get_task_description_display(name): #adited
    return Task_Description_Display_Map.get(name, name.replace('_', ' '))  



def normalize_project_key(name):
    return name.strip().replace('_', ' ').lower()


def get_project_display(name):
    return PROJECT_DISPLAY_MAP.get(name, name.replace('_', ' '))


def week_number_to_month(week_number):
    for month_index, max_week in enumerate(WEEK_MONTH_THRESHOLDS, start=1):
        if week_number <= max_week:
            return month_index
    return 12


# Create your views here.

class HomePageView(TemplateView):
    template_name = 'home.html'

class AboutPageView(TemplateView):
    template_name = 'about.html'

@login_required(login_url='login')
def dashboard_view(request):
    form = NewTimesheetForm(request.POST or None)
    message = None
    if request.method == 'POST' and form.is_valid():
        employee = form.cleaned_data['employee'].strip()
        project_name = form.cleaned_data['project_name'].strip()
        week_number = form.cleaned_data.get('week_number')
        hours_week = form.cleaned_data['hours_week']

        last_week = Timesheet.objects.filter(employee__iexact=employee).aggregate(
            Max('week_number')
        )['week_number__max'] or 0

        if week_number is None:
            if last_week >= 52:
                message = 'This employee already has hours logged for all 52 weeks.'
            else:
                week_number = last_week + 1
        else:
            if week_number < 1 or week_number > 52:
                message = 'Week number must be between 1 and 52.'

        if message is None:
            Timesheet.objects.create(
                employee=employee,
                project_name=project_name,
                task_description=form.cleaned_data['task_description'],
                week_number=week_number,
                hours_week=hours_week,
            )
            return redirect('dashboard')

    selected_employee = request.GET.get('employee', 'all')
    start_month = int(request.GET.get('start_month', 1))
    end_month = int(request.GET.get('end_month', 12))
    if start_month < 1:
        start_month = 1
    if end_month > 12:
        end_month = 12
    if end_month < start_month:
        end_month = start_month

    employee_choices = ['all'] + list(
        Timesheet.objects.order_by('employee')
        .values_list('employee', flat=True)
        .distinct()
    )

#FIXES
    entries = Timesheet.objects.all()
    if request.GET.get('employee', 'all') != 'all':
        entries = entries.filter(employee__iexact=request.GET['employee'])

    project_display = {}
    project_order = []
    for raw_project in entries.values_list('project_name', flat=True):
        key = normalize_project_key(raw_project)
        if key not in project_display:
            project_display[key] = get_project_display(raw_project)
            project_order.append(key)

    month_hours = {project_key: [0] * 12 for project_key in project_order}

    for entry in entries:
        project_key = normalize_project_key(entry.project_name)
        if project_key not in month_hours:
            project_display[project_key] = get_project_display(entry.project_name)
            month_hours[project_key] = [0] * 12
            project_order.append(project_key)

        month_index = week_number_to_month(entry.week_number) - 1
        if 0 <= month_index < 12:
            month_hours[project_key][month_index] += float(entry.hours_week)

    months = MONTH_NAMES[start_month - 1:end_month]

    datasets = []
    for project_key in project_order:
        data = month_hours[project_key][start_month - 1:end_month]
        if any(value != 0 for value in data):
            datasets.append({
                'label': project_display[project_key],
                'data': data,
                'backgroundColor': f'rgba({hash(project_key) % 256}, {(hash(project_key) // 256) % 256}, {(hash(project_key) // 65536) % 256}, 0.6)',
                'borderColor': 'transparent',
                'borderWidth': 1,
            })

    # --- PLANNED HOURS LINES PER PROJECT ---
    # NOTE: Planned-hours lines intentionally removed per user request.

    chart_data = {
        'labels': months,
        'datasets': datasets,
    }

    table_data = {}
    for entry in Timesheet.objects.all():
        row = table_data.setdefault(entry.employee, [0] * 52)
        if 1 <= entry.week_number <= 52:
            row[entry.week_number - 1] += float(entry.hours_week)
    for employee, weeks in table_data.items():
        table_data[employee] = [hours if hours != 0 else None for hours in weeks]

    # --- PIE CHART DATA (by task description) ---
    pie_data = {}
    pie_metadata = {}  # stores employee and project info for tooltips
    
    pie_entries = Timesheet.objects.all()
    if request.GET.get('employee', 'all') != 'all':
        pie_entries = pie_entries.filter(employee__iexact=request.GET['employee'])
    
    for entry in pie_entries:
        task_key = get_task_description_display(entry.task_description)
        project_key = normalize_project_key(entry.project_name)
        project_name = project_display.get(project_key, get_project_display(entry.project_name))
        
        if task_key not in pie_data:
            pie_data[task_key] = 0
            pie_metadata[task_key] = {'employees': set(), 'project': project_name}
        
        pie_data[task_key] += float(entry.hours_week)
        pie_metadata[task_key]['employees'].add(entry.employee)
    
    # Create color mapping for projects (matching bar chart colors)
    task_colors = []
    task_projects = []
    for task_key in pie_data.keys():
        project_name = pie_metadata[task_key]['project']
        project_key = normalize_project_key(project_name)
        task_projects.append(project_name)
        # Use same color generation as bar chart
        color = f"rgba({hash(project_key) % 256}, {(hash(project_key) // 256) % 256}, {(hash(project_key) // 65536) % 256}, 0.7)"
        task_colors.append(color)
    
    pie_chart_data = {
        'labels': list(pie_data.keys()),
        'datasets': [{
            'data': list(pie_data.values()),
            'backgroundColor': task_colors,
            'borderColor': '#fff',
            'borderWidth': 2,
        }],
        'metadata': {k: {'employees': list(v['employees']), 'project': v['project']} for k, v in pie_metadata.items()}
    }

    return render(
        request,
        'dashboard.html',
        {
            'form': form,
            'message': message,
            'table_data': table_data,
            'week_headers': range(1, 53),
            'chart_data_json': json.dumps(chart_data),
            'pie_chart_json': json.dumps(pie_chart_data),
            'employee_choices': employee_choices,
            'selected_employee': selected_employee,
            'start_month': start_month,
            'end_month': end_month,
            'month_choices': list(enumerate(MONTH_NAMES, start=1)),
        },
    )

@login_required(login_url='login')
def generate_report(request):
    """Export all timesheet entries to a fixed XLSX report file."""

    entries = Timesheet.objects.all().order_by('week_number', 'employee')
    report_path = Path(r"C:\Users\Senzo Mafu\OneDrive - Elwatini RC\Desktop\TimeTrack\Timetrack_Report.xlsx")
    report_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    summary = workbook.active
    summary.title = 'Summary'

    summary.append(['TimeTrack Report'])
    summary.append(['Generated by', request.user.username])
    summary.append(['Total Timesheet Rows', entries.count()])
    summary.append(['Total Hours', sum(float(e.hours_week) for e in entries)])
    summary.append([])

    summary.append(['Hours by Project'])
    summary.append(['Project', 'Hours'])
    hours_by_project = {}
    for e in entries:
        project = get_project_display(e.project_name)
        hours_by_project[project] = hours_by_project.get(project, 0) + float(e.hours_week)
    for project, hours in sorted(hours_by_project.items()):
        summary.append([project, float(f"{hours:.2f}")])
    summary.append([])

    summary.append(['Hours by Employee'])
    summary.append(['Employee', 'Hours'])
    hours_by_employee = {}
    for e in entries:
        hours_by_employee[e.employee] = hours_by_employee.get(e.employee, 0) + float(e.hours_week)
    for employee, hours in sorted(hours_by_employee.items()):
        summary.append([employee, float(f"{hours:.2f}")])
    summary.append([])

    summary.append(['Hours by Task'])
    summary.append(['Task', 'Hours'])
    hours_by_task = {}
    for e in entries:
        task = get_task_description_display(e.task_description)
        hours_by_task[task] = hours_by_task.get(task, 0) + float(e.hours_week)
    for task, hours in sorted(hours_by_task.items()):
        summary.append([task, float(f"{hours:.2f}")])

    employee_monthly_summary = workbook.create_sheet(title='Employee Monthly Summary')
    employee_monthly_summary.append(['Employee', *MONTH_NAMES, 'Total'])
    employee_month_totals = {}
    for e in entries:
        employee = e.employee
        month_name = MONTH_NAMES[week_number_to_month(e.week_number) - 1]
        if employee not in employee_month_totals:
            employee_month_totals[employee] = {'months': {month: 0.0 for month in MONTH_NAMES}, 'total': 0.0}
        employee_month_totals[employee]['months'][month_name] += float(e.hours_week)
        employee_month_totals[employee]['total'] += float(e.hours_week)
    for employee in sorted(employee_month_totals):
        month_values = employee_month_totals[employee]['months']
        employee_monthly_summary.append([
            employee,
            *[float(f"{month_values[month]:.2f}") for month in MONTH_NAMES],
            float(f"{employee_month_totals[employee]['total']:.2f}")
        ])

    details = workbook.create_sheet(title='Entries')
    details.append(['Employee', 'Project', 'Task Description', 'Week Number', 'Month', 'Hours'])
    for e in entries:
        details.append([
            e.employee,
            get_project_display(e.project_name),
            get_task_description_display(e.task_description),
            e.week_number,
            MONTH_NAMES[week_number_to_month(e.week_number) - 1],
            float(e.hours_week)
        ])

    workbook.save(report_path)

    with report_path.open('rb') as excel_file:
        response = HttpResponse(
            excel_file.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{report_path.name}"'
        return response



class LoginPageView(LoginView):
    template_name = 'login.html'
    success_url = reverse_lazy('dashboard')

class LogoutPageView(LogoutView):
    next_page = reverse_lazy('home')

    def get_success_url(self):
        return reverse_lazy('home')

class TimesheetUpdateView(UpdateView):
    model = Timesheet
    form_class = TimesheetForm
    template_name = 'edit_timesheet.html'
    success_url = reverse_lazy('dashboard')

class TimesheetDeleteView(DeleteView):
    model = Timesheet
    template_name = 'delete_timesheet.html'
    success_url = reverse_lazy('dashboard')



